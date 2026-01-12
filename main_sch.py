# main_sch.py
 
from schemes import scrape_schemes
from config import COMPANIES
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# Timestamped output
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
OUTPUT_FILE = Path(f"output/test_schemes_{timestamp}.xlsx")
OUTPUT_FILE.parent.mkdir(exist_ok=True)
 

for company in COMPANIES:
    print(f"Scraping schemes for {company}...")
    schemes_df = scrape_schemes(company)

    if schemes_df.empty:
        print(f"No data for {company}")
        continue

    sheet_name = company[:31]

    if OUTPUT_FILE.exists():
        with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            if sheet_name in writer.book.sheetnames:
                start_row = writer.book[sheet_name].max_row + 3
            else:
                start_row = 0

            schemes_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=start_row,
                header=(start_row == 0)
            )
    else:
        with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl", mode="w") as writer:
            schemes_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

print(f"✅ Schemes Excel generated: {OUTPUT_FILE}")
 
 